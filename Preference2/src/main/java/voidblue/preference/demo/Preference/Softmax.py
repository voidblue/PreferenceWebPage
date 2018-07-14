import tensorflow as tf
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import os
#Softmax 단일변수는 의미가 없기때문에 다중 변수만 사용 가능
#입력값은 x[i]가 통채로 하나씩 들어간다.
#x_data의 shape는  (입력데이터의 집합개수 , feature의 개수)
#y_data의 shape는 (feature의 개수 , 신경의 수)
#y_data학습할때는 반드시 0또는 1이여야 한다.
path = os.path.dirname(os.path.realpath(__file__))
class Softmax:
    layer = 0
    x_data = None
    y_data = None
    W_val = []
    cost_val = []
    X_val = []
    Y_val = []
    X = None
    Y = None
    layers = []
    weights = []
    bias = []
    hypothesis = None
    cost = None
    learning_rate=None
    sess= None
    result = None
    layer = 0
    list_step = 0

    def __init__(self,x_data,y_data):
        self.x_data = x_data
        self.y_data = y_data
        self.layers = []
        self.weights = []
        self.bias = []

        self.X = tf.placeholder(tf.float32,[None, None])
        self.Y = tf.placeholder(tf.float32, [None , None])
        self.sess = tf.Session()


    #코스트를 만드는데 필요한 weight, bias, hypothesis 모두 정의 함수이름 생각나면 변경하기
    def set_cost(self,input_data,input_length):
        # self.x_data = input_data
        self.bias.append(tf.Variable(tf.random_uniform([len(self.y_data[0])], -1.0, 1.0)+10,dtype=tf.float32))
        # self.weights.append(tf.Variable(tf.random_uniform([input_length,len(self.y_data[0])], -1.0, 1.0),dtype=tf.float32))
        i = 0
        while(True):
            try:
                self.weights.append(tf.get_variable("weight"+str(i), shape=[input_length, len(self.y_data[0])],
                initializer=tf.contrib.layers.xavier_initializer()))

                break
            except:
                pass
            i+=1
        # init = tf.global_variables_initializer()
        # self.sess.run(init)

        # self.bias.append(tf.Variable(tf.random_uniform([tf.size(self.y_data[0])], -1.0, 1.0)))
        # self.weights.append(tf.Variable(tf.random_uniform([input_length, tf.size(self.y_data[0])], -1.0, 1.0)))

        self.hypothesis = tf.nn.softmax(tf.add(tf.matmul(input_data,-self.weights[-1]),self.bias[-1]))
        self.hypothesis = tf.clip_by_value(self.hypothesis, 1e-25, 1-(1e-25))
        self.cost = tf.reduce_mean(-tf.reduce_sum(self.Y*tf.log(self.hypothesis),reduction_indices=1))
        # self.cost = -tf.reduce_mean(self.y_data * tf.log(self.hypothesis)) \
        #             + tf.subtract(1., self.y_data) * tf.log(tf.subtract(1., self.hypothesis))
        # init = tf.initialize_all_variables()
        # self.sess.run(init)

    #학습 시작
    def training(self, learning_rate=0.0004, step=200, show_training_data=False, cost_limit = 0, mode = 'adadelta', dropout = 1):
        if mode == 'adadelta':
            train = tf.train.AdadeltaOptimizer(learning_rate=learning_rate).minimize(self.cost)
        elif mode == 'adam':
            train = tf.train.AdamOptimizer(learning_rate=learning_rate).minimize(self.cost)
        elif mode == 'momentum':
            train = tf.train.MomentumOptimizer(learning_rate=learning_rate, momentum=0.9).minimize(self.cost)
        elif mode == 'adagrad':
            train = tf.train.AdagradOptimizer(learning_rate=learning_rate).minimize(self.cost)
        init = tf.global_variables_initializer()
        self.sess.run(init)

        self.list_step =range(step)
        for step in range(step):
            try:
                self.sess.run(train, feed_dict={self.X: self.x_data, self.Y: self.y_data})
                cost = self.sess.run(self.cost, feed_dict={self.X: self.x_data, self.Y: self.y_data})
            except:
                self.sess.run(train, feed_dict={self.X: self.sess.run(self.x_data), self.Y: self.y_data})
                cost = self.sess.run(self.cost, feed_dict={self.X: self.sess.run(self.x_data), self.Y: self.y_data})

            if show_training_data==True and step % 100 == 0 :
                try:
                    print('weght = ',self.sess.run(self.weights,feed_dict={self.X: self.x_data, self.Y: self.y_data}),'hyp = ',self.sess.run(self.hypothesis,feed_dict={self.X: self.x_data, self.Y: self.y_data}),'cost =',cost)
                except:
                    print('weght = ',
                          self.sess.run(self.weights, feed_dict={self.X: self.sess.run(self.x_data), self.Y: self.y_data}), 'hyp = ',
                          self.sess.run(self.hypothesis, feed_dict={self.X: self.sess.run(self.x_data), self.Y: self.y_data}),
                          'cost =', cost)
            elif(step % 100 == 0):
                print(str(step) + " cost : " + str(cost))
            if cost < cost_limit:
                break
            # self.cost_val.append(cost)
    def show_cost_graph(self):
        plt.plot(self.list_step, self.cost_val, 'ro')
        plt.ylabel('cost')
        plt.xlabel('step')
        plt.show()

    def save_graphToExcel(self, filename, inputfragment, outputdic, numOfOutput = 1):
        outlength = len(outputdic)
        result = []
        # result = self.sess.run(tf.argmax(self.predict(self.x_data), 1))
        temp = list(self.predict(self.x_data))
        for i, item in enumerate(temp):
            temp[i] = list(item)
        for item in temp:
            for x in range(numOfOutput):
                temp2 = item.index(max(item))
                result.append(int(temp2))
                item[item.index(max(item))] = 0

        # print('길이',len(result))
        # 엑셀파일 열기
        try:
            wb = openpyxl.load_workbook(filename + '.xlsx')
        except:
            wb = openpyxl.Workbook()
            wb.save(filename + ".xlsx")

        # 현재 Active Sheet 얻기
        ws = wb.active
        # ws = wb.get_sheet_by_name("Sheet1")
        inputset = set()
        for item in inputfragment:
            inputset.add(item)
        for r in range(outlength):
            ws.cell(row=r+1, column=1).value = outputdic[56 - r]

        inputsetlist = list(inputset)
        inputsetlist.sort()
        ws.cell(row=outlength+1, column=1).value = "여행지/입력번호"
        for j in range(len(inputsetlist)):
            ws.cell(row = outlength +1, column=j+2).value = inputsetlist[j]
        for j in range(len(inputsetlist)):
            for k in range(outlength):
                ws.cell(row=k+1 , column = j+2).value = 0

        for j in range(len(inputfragment)):
            for k in range(numOfOutput):
                ws.cell(row=outlength - result[j+k],
                        column= np.where(inputsetlist == inputfragment[j])[0][0] +2 ).value += 1
        # 엑셀 파일 저장"

        wb.save(filename + ".xlsx")
        wb.close()


    def show_graph(self, xlabel, inputfragment):
        dataset = set()
        size = []
        result = self.sess.run(tf.argmax(self.predict(self.x_data), 1))
        dataset = set()
        dic = {}
        for i in range(len(inputfragment)):
            before = len(dataset)

            dataset.add(str(inputfragment[i]) + 'ESC' + str(result[i]))
            after = len(dataset)

            if after - before == 1:
                dic[str(inputfragment[i]) + 'ESC' + str(result[i])] = 3
            else:
                dic[str(inputfragment[i]) + 'ESC' + str(result[i])] += 3

        for i in range(len(inputfragment)):
            size.append(dic[str(inputfragment[i]) + 'ESC' + str(result[i])]**0.5)
        plt.scatter(inputfragment, result, s=size)

        plt.show()

    #조건에 맞는 입력 데이타를 받으면 회귀에 따라 예측이되는 출력값을 보낸다
    def predict(self, x_data):
        temp = x_data
        x = 0
        for i in range(len(self.weights)-1):
            temp = tf.add(tf.matmul(temp,-self.weights[i]), self.bias[i])
        self.hypothesis = tf.nn.softmax(tf.add(tf.matmul(temp, -self.weights[-1]), self.bias[-1]))
        #딥하게 가면 입력값이 텐서가 되고 아니면 리스트이기 때문에 처리해준다.

        try:
            self.result = self.sess.run(self.hypothesis, feed_dict={self.X:self.sess.run(temp)})
        except:
            self.result = self.sess.run(self.hypothesis, feed_dict={self.X:temp})

        # print(self.sess.run(tf.argmax(self.result,1)) , "원핫 인코딩 후 값") #argmax가 one-hot encoding
        return self.result

    #one-hot encoding하기 전에 소프트맥스 출력으로 나오는 상대적 비율을 받는다.
    def return_predict_possibility(self):
        return self.result

    def return_predict_onehot(self):
        return self.sess.run(tf.argmax(self.result, 1))


    def save_weight(self, weightname = 'weight', biasname = 'bias'):
        for i in range (len(self.weights)):
            np.save(weightname+"layer"+str(i), self.sess.run(self.weights[i]))
        # np.save(weightname, self.sess.run(self.weights))
        np.save(biasname, self.sess.run(self.bias))


        # except:
        #     np.save(weightname,self.sess.run(self.weights[0]))
        #     np.save(biasname,self.sess.run(self.bias[0]))


    def load_weight(self, layer, weightname = 'weight', biasname = 'bias'):
        self.weights = [0] * layer
        for i in range(layer):
            self.weights[i] = np.load(path +"/"+ weightname+ "layer" + str(i)+".npy")
        self.bias = np.load(path + "/" + biasname+".npy")

    #딥하게 층을 하나 더 만든다. X는 리스트가 올수도 있고, 텐서가 올수도 있다.
    def create_layer(self, X, input_length, output_length, dropout = 1):
        self.weights.append(tf.Variable(tf.random_uniform([input_length, output_length],-1.0,1.0)))
        self.bias.append(tf.Variable(tf.random_uniform([output_length], -1.0, 1.0)))

        ret = tf.add(tf.matmul(X,-self.weights[-1]),self.bias[-1])
        # ret = tf.nn.relu(ret)
        self.layers.append(ret)
        return ret


