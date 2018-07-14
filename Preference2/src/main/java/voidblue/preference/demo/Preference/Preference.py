import numpy as np
import openpyxl
import tensorflow as tf
import numpy.core._methods
import numpy.lib.format
from Softmax import Softmax
import os
import sys

if getattr(sys, 'frozen', False):
    # frozen
    dir_ = os.path.dirname(sys.executable)
else:
    # unfrozen
    dir_ = os.path.dirname(os.path.realpath(__file__))

class Pref(Softmax):



    def process_ranking(self, target_data = None ):
        xlist = []
        first = 0
        second = 0
        third = 0
        # print(target_data)
        for y in target_data:
            first = 0
            second = 0
            third = 0
            forth = 0
            fifth = 0
            for item in y:
                # if xxx > third:
                #     third = xxx
                # if third > second:
                #     temp = second
                #     second = third
                #     third = temp
                # if second > first:
                #     temp = first
                #     first = second
                #     second = temp
                #
                # if item > first:
                #     first = item

                # if third > second:
                #     temp = second
                #     second = third
                #     third = temp
                # if second > first:
                #     temp = first
                #     first = second
                #     second = temp

                if item > fifth:
                    fifth = item
                if fifth > forth:
                    temp = forth
                    forth = fifth
                    fifth = temp
                if forth > third:
                    temp = third
                    third = forth
                    forth = temp
                if third > second:
                    temp = second
                    second = third
                    third = temp
                if second > first:
                    temp = first
                    first = second
                    second = temp

            if len(list(np.where(y == first))) >= 1:
                xlist.append(list(np.where(y == first)[0]))
            else:
                xlist.append([0])
            if len(list(np.where(y == second))) >= 1:
                xlist.append(list(np.where(y == second)[0]))
            else:
                xlist.append([0])
            if len(list(np.where(y == third))) >= 1:
                xlist.append(list(np.where(y == third)[0]))
            else:
                xlist.append([0])
            xlist.append(list(np.where(y == forth)[0]))
            xlist.append(list(np.where(y == fifth)[0]))
        return list(xlist)

    def target_ranking(self, x_data):
        temp = x_data
        for i, w in enumerate(self.weights):
            temp = tf.nn.softmax(tf.add(tf.matmul(temp,-w),self.bias[i]))
        return self.sess.run(temp)

    def return_rank_accuracy(self, y_data ,data):
        try:
            wb = openpyxl.load_workbook("테스트결과" + '.xlsx')
        except:
            wb = openpyxl.Workbook()
            wb.save("테스트결과" + ".xlsx")
        ws = wb.active
        accuracy = 0
        sametime = [0, 0, 0, 0, 0]
        for i in range(len(y_data)):
            acc = 0
            ws.cell(row=i +2, column= 14).value = 1
            if y_data[i][data[5*i]] == 1:
                accuracy += 1
                sametime[0]+=1
                acc += 1
            elif y_data[i][data[(5 * i) + 1]] == 1:
                accuracy += 1
                sametime[1] += 1
                acc += 1
            elif y_data[i][data[(5 * i) + 2]] == 1:
                accuracy += 1
                sametime[2] += 1
                acc += 1
            elif y_data[i][data[(5 * i) + 3]] == 1:
                accuracy += 1
                sametime[3] += 1
                acc += 1
            elif y_data[i][data[(5 * i) + 4]] == 1:
                accuracy += 1
                sametime[4] += 1
                acc += 1
            else:
                ws.cell(row=i + 2, column= 14).value = 0
            
            # if(acc == 3):
            #     print(i, "번쨰 데이터가 모두 적중 (0부터 시작)")
            #     break
        # print(sametime)
        wb.save("테스트결과" + '.xlsx')
        wb.close()
        accuracy /= len(y_data)
        print('accuracy = {}'.format(accuracy))
