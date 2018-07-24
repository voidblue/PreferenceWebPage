import openpyxl

import preprocess
from DBGetter import DBGetter
from LoadData import LoadData
from Preference import Pref

place = {1:"한라산@@@자연-숲", 2:"오름@@@자연-숲", 3: "성산일출봉@@@자연-숲", 4 : "섬(우도/마라도 등)@@@자연명소",
         5:"올레길@@@기타", 6: "폭포(정방폭포 등)@@@자연명소", 7: "동굴(만장굴 등)@@@자연명소", 8: "해수욕장@@@자연명소",
         9:"비자림@@@자연-숲",  10:"한라수목원@@@공원", 11: "서귀포자연휴양림@@@자연-숲", 12: "절물자연휴양림@@@자연-숲",
         13:"용두암@@@자연명소", 14:"주상절리대@@@자연명소", 15: "한림공원@@@공원", 16:"국립제주박물관@@@박물관", 17:"도립미술관@@@박물관", 18:"민속자연사박물관@@@박물관",
         19:"제주돌문화공원@@@공원", 20:"제주세계자연유산센터@@@박물관", 21:"이중섭미술관@@@박물관", 22:"서복전시관@@@박물관", 23:"제주4・3평화공원@@@공원",
         24:"동문시장@@@쇼핑", 25:"중앙로지하상가@@@쇼핑", 26:"바오젠거리@@@쇼핑",27:"제주5일장(제주/서귀포)@@@쇼핑",
         28:"서귀포매일올레시장@@@쇼핑", 29:"신라면세점@@@면세점", 30:"롯데면세점@@@면세점", 31:"제주관광공사 면세점@@@면세점", 32: "공항JDC면세점@@@면세점",
         33:"제주목관아@@@문화유적", 34:"항몽유적지@@@문화유적", 35:"성읍민속마을@@@문화유적",
         36:"삼양동선사유적@@@문화유적", 37:"제주추사관@@@문화유적", 38:"관덕정@@@문화유적", 39:"이중섭거주지@@@문화유적", 40:"하멜기념비@@@문화유적",
         41:"미로공원(김녕/메이지)@@@테마파크", 42:"에코랜드@@@테마파크", 43:"제주경마공원@@@체험", 44:"불교사찰@@@기타",
         45:"아쿠아플라넷@@@체험", 46:"테디베어박물관@@@테마파크", 47:"소인국테마파크@@@테마파크", 48:"잠수함관광(서귀포 등)@@@체험",
         49:"신비의도로@@@기타", 50:"생각하는정원@@@공원", 51:"드라마촬영지@@@기타", 52:"제주별빛누리공원@@@테마파크", 53:"유람선(요트투어포함)@@@체험",
        54:"제주바다체험장@@@체험", 55:"골프장@@@기타", 56: "카지노@@@기타"}

def printAccuracy():
    dbGetter = DBGetter()
    dbGetter.connect()
    origin_x_data = dbGetter.getInput()
    y_data = dbGetter.getOutput()
    dbGetter.close()
    recommendvalue, testy_data = merge(origin_x_data, y_data)

    try:
        wb = openpyxl.load_workbook("테스트결과" + '.xlsx')
    except:
        wb = openpyxl.Workbook()
        wb.save("테스트결과" + ".xlsx")

    # 현재 Active Sheet 얻기
    ws = wb.active
    for i, w in enumerate(recommendvalue):
        if i % 5 == 0:
            for j in range(5):
                ws.cell(row=i//5 + 2, column =j + 1).value = place[recommendvalue[i+j] +1]
        if i % 5 == 0:
            k = 0
            for j, test in enumerate(testy_data[i//5]):
                if (test == 1):
                    ws.cell(row=i // 5 + 2, column=k + 7).value = place[j + 1]
                    k+=1

    wb.save("테스트결과" + ".xlsx")
    wb.close()
    Pref(None, None).return_rank_accuracy(testy_data, recommendvalue)

def recomend(x_param):
    print("???뭐죠")
    dbGetter = DBGetter()
    dbGetter.connect()
    origin_x_data = [x_param]
    y_data = [dbGetter.getOutput()[0]]

    less_input_data = []

    origin_x_data = preprocess.transpose(origin_x_data)
    less_input_data.append(list(origin_x_data[0]))
    less_input_data.append(list(origin_x_data[1]))
    less_input_data.append(list(origin_x_data[2]))
    less_input_data.append(list(origin_x_data[3]))
    less_input_data.append(list(origin_x_data[5]))
    less_input_data.append(list(origin_x_data[13]))
    less_input_data.append(list(origin_x_data[23]))
    less_input_data.append(list(origin_x_data[24]))
    less_input_data.append(list(origin_x_data[25]))
    less_input_data.append(list(origin_x_data[26]))
    less_input_data.append(list(origin_x_data[27]))
    less_input_data.append(list(origin_x_data[28]))
    less_input_data.append(list(origin_x_data[29]))
    less_input_data.append(list(origin_x_data[30]))

    less_input_data = preprocess.transpose(less_input_data)

    input = less_input_data
    expected = y_data
    x_data = dbGetter.getInput()
    x_data = preprocess.transpose(x_data)
    input = preprocess.transpose(input)
    input = preprocess.zerotoone(x_data, input)
    input = preprocess.transpose(input)

    preference = Pref(input, expected)
    feat1 = [0] * 56
    result = []
    for i in range(len(input)):
        result.append(list(feat1))
    for i in range(0, 9):
        preference.load_weight(5, 'weight' + str(i), 'bias' + str(i))
        data = preference.predict(list(input))
        rank = preference.process_ranking(data)
        for j, w in enumerate(rank):
            if j % 5 == 0:
                result[j // 5][w[0]] += 5
            if j % 5 == 1:
                result[j // 5][w[0]] += 4
            if j % 5 == 2:
                result[j // 5][w[0]] += 3
            if j % 5 == 3:
                result[j // 5][w[0]] += 2
            if j % 5 == 4:
                result[j // 5][w[0]] += 1

        preference.xlist = []

    finresult = []
    for y in result:
        first = 0
        second = 0
        third = 0
        forth = 0
        fifth = 0

        p1 = 0
        p2 = 0
        p3 = 0
        p4 = 0
        p5 = 0
        for i, item in enumerate(y):

            if item > fifth:
                fifth = item
                p5 = i
            if fifth > forth:
                temp = int(forth)
                forth = int(fifth)
                fifth = temp
                p5 = int(p4)
                p4 = i
            if forth > third:
                temp = int(third)
                third = int(forth)
                forth = temp
                p4 = int(p3)
                p3 = i
            if third > second:
                temp = int(second)
                second = int(third)
                third = temp
                p3 = int(p2)
                p2 = i
            if second > first:
                temp = int(first)
                first = int(second)
                second = temp
                p2 = int(p1)
                p1 = i
        finresult.append(p1)
        finresult.append(p2)
        finresult.append(p3)
        finresult.append(p4)
        finresult.append(p5)
    print("???뭐죠")
    print(end='ESC')
    print(place[finresult[0]+1],end='ESC')
    print(place[finresult[1]+1],end='ESC')
    print(place[finresult[2]+1],end='ESC')
    print(place[finresult[3]+1],end='ESC')
    print(place[finresult[4]+1],end='ESC')
    print("")
    dbGetter.close()
def merge(origin_x_data, y_data):

    less_input_data = []

    origin_x_data = preprocess.transpose(origin_x_data)
    less_input_data.append(list(origin_x_data[0]))
    less_input_data.append(list(origin_x_data[1]))
    less_input_data.append(list(origin_x_data[2]))
    less_input_data.append(list(origin_x_data[3]))
    less_input_data.append(list(origin_x_data[5]))
    less_input_data.append(list(origin_x_data[13]))
    less_input_data.append(list(origin_x_data[23]))
    less_input_data.append(list(origin_x_data[24]))
    less_input_data.append(list(origin_x_data[25]))
    less_input_data.append(list(origin_x_data[26]))
    less_input_data.append(list(origin_x_data[27]))
    less_input_data.append(list(origin_x_data[28]))
    less_input_data.append(list(origin_x_data[29]))
    less_input_data.append(list(origin_x_data[30]))

    less_input_data = preprocess.transpose(less_input_data)
    testx_data = []
    testy_data = []
    for j in range(len(y_data)):
        if sum(y_data[j]) == 5:
            testx_data.append(list(less_input_data[j]))
            testy_data.append(list(y_data[j]))

    x_data = preprocess.transpose(testx_data)
    testx_data = preprocess.transpose(testx_data)
    testx_data = preprocess.zerotoone(x_data, testx_data)
    testx_data = preprocess.transpose(testx_data)
    x_data = preprocess.zerotoone(x_data, x_data)
    x_data = preprocess.transpose(x_data)

    preference = Pref(testx_data, testy_data)
    # layer1 = preference.create_layer(x_data, 31, 31)
    # layer2 = preference.create_layer(layer1, 56, 56)
    # preference.show_cost_graph()
    weight = None
    feat1 = [0] * 56
    result = []
    for i in range(len(testx_data)):
        result.append(list(feat1))
    for i in range(0, 9):
        preference.load_weight(5, 'weight' + str(i), 'bias' + str(i))
        data = preference.predict(list(testx_data))
        rank = preference.process_ranking(data)
        for j, w in enumerate(rank):
            if j % 5 == 0:
                result[j // 5][w[0]] += 5
            if j % 5 == 1:
                result[j // 5][w[0]] += 4
            if j % 5 == 2:
                result[j // 5][w[0]] += 3
            if j % 5 == 3:
                result[j // 5][w[0]] += 2
            if j % 5 == 4:
                result[j // 5][w[0]] += 1

        preference.xlist = []
    # print(result)

    xlist = []
    finresult = []
    for y in result:
        first = 0
        second = 0
        third = 0
        forth = 0
        fifth = 0

        p1 = 0
        p2 = 0
        p3 = 0
        p4 = 0
        p5 = 0
        for i, item in enumerate(y):

            if item > fifth:
                fifth = item
                p5 = i
            if fifth > forth:
                temp = int(forth)
                forth = int(fifth)
                fifth = temp
                p5 = int(p4)
                p4 = i
            if forth > third:
                temp = int(third)
                third = int(forth)
                forth = temp
                p4 = int(p3)
                p3 = i
            if third > second:
                temp = int(second)
                second = int(third)
                third = temp
                p3 = int(p2)
                p2 = i
            if second > first:
                temp = int(first)
                first = int(second)
                second = temp
                p2 = int(p1)
                p1 = i
        finresult.append(p1)
        finresult.append(p2)
        finresult.append(p3)
        finresult.append(p4)
        finresult.append(p5)
    return finresult , testy_data








